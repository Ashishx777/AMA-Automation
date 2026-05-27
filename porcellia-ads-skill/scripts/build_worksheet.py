#!/usr/bin/env python3
"""Build a copywriter-friendly worksheet from an images folder.

Output: an .xlsx where column A is a thumbnail of each image and the
adjacent columns are empty copy fields the copywriter fills in. The
copywriter SEES the photo they're writing for — no descriptions, no
guessing, no cross-referencing.

Usage::

    python build_worksheet.py \\
        --images   D:/path/to/images \\
        --archetype ama \\
        --out      D:/path/to/worksheet.xlsx \\
        [--pre-fill D:/path/to/existing-copy.csv]   # optional

The output .xlsx has:
    Column A: image thumbnail (~200px wide, 200px row height)
    Column B: id              (auto-filled: v1, v2, …)
    Column C: image_filename  (the actual filename, used by the pipeline)
    Column D…: archetype-specific copy columns

When the copywriter finishes, they "Save As CSV" or "Export to CSV"
inside Excel/Sheets. That CSV drops straight into the pipeline.
"""
from __future__ import annotations

import argparse
import csv as csv_mod
import io
import sys
from pathlib import Path

try:
    from PIL import Image
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError as e:
    raise SystemExit(f"Missing dependency: {e}. Run: pip install pillow openpyxl")


ARCHETYPE_COLUMNS = {
    "ama":   ["context_bar", "question_text", "answer_text", "image_note"],
    "forum": ["community", "handle", "title", "body", "upvotes", "comments"],
}

IMAGE_EXTS = {".jpg", ".jpeg", ".png", ".webp", ".gif"}
THUMB_PX = 200
ROW_HEIGHT_PX = 200


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--images", required=True, type=Path,
                    help="Folder containing source images")
    ap.add_argument("--archetype", default="ama",
                    choices=list(ARCHETYPE_COLUMNS))
    ap.add_argument("--out", required=True, type=Path,
                    help="Output .xlsx path")
    ap.add_argument("--pre-fill", type=Path, default=None,
                    help="Optional CSV to pre-populate copy columns (matched by id)")
    args = ap.parse_args()

    if not args.images.exists():
        raise SystemExit(f"Images folder not found: {args.images}")

    # Sort by numeric stem when possible (v1, v2, …, v10 in correct order)
    images = sorted(
        [p for p in args.images.iterdir()
         if p.is_file() and p.suffix.lower() in IMAGE_EXTS],
        key=_natural_key,
    )
    if not images:
        raise SystemExit(f"No images found in {args.images}")

    pre_fill = _load_prefill(args.pre_fill, args.archetype) if args.pre_fill else {}

    wb = Workbook()
    ws = wb.active
    ws.title = "Copy worksheet"

    # ── Header row ─────────────────────────────────────────────────────────
    columns = ["image", "id", "image_filename"] + ARCHETYPE_COLUMNS[args.archetype]
    for col_idx, name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1f1f23")
        cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 28

    # Column widths
    ws.column_dimensions["A"].width = 30   # image thumb
    ws.column_dimensions["B"].width = 8    # id
    ws.column_dimensions["C"].width = 22   # image_filename
    for i, _ in enumerate(ARCHETYPE_COLUMNS[args.archetype]):
        col_letter = get_column_letter(4 + i)
        ws.column_dimensions[col_letter].width = 40

    # ── One row per image ──────────────────────────────────────────────────
    tmp_paths: list[Path] = []
    for i, src in enumerate(images, start=1):
        row_num = i + 1
        vid = f"v{i}"
        thumb_path = _make_thumb(src, THUMB_PX)
        tmp_paths.append(thumb_path)

        # Insert image
        xlimg = XLImage(str(thumb_path))
        xlimg.width = THUMB_PX
        xlimg.height = THUMB_PX
        ws.add_image(xlimg, f"A{row_num}")
        ws.row_dimensions[row_num].height = ROW_HEIGHT_PX * 0.75  # excel uses pt

        # id
        ws.cell(row=row_num, column=2, value=vid).alignment = Alignment(vertical="center")
        # image_filename
        ws.cell(row=row_num, column=3, value=src.name).alignment = Alignment(vertical="center", wrap_text=True)

        # Empty copy columns (or pre-filled if we have a pre-fill CSV)
        for j, colname in enumerate(ARCHETYPE_COLUMNS[args.archetype]):
            v = pre_fill.get(vid, {}).get(colname, "")
            c = ws.cell(row=row_num, column=4 + j, value=v)
            c.alignment = Alignment(vertical="center", wrap_text=True)

    # Freeze the header row.
    ws.freeze_panes = "A2"

    # Instructions sheet
    notes = wb.create_sheet("Read me")
    notes_lines = [
        ("How to fill this worksheet", True),
        ("", False),
        ("1. Each row already has the image you're writing copy for.", False),
        ("2. Don't touch column A (image), B (id), or C (image_filename).", False),
        (f"3. Fill in: {', '.join(ARCHETYPE_COLUMNS[args.archetype])}.", False),
        ("4. When finished: File → Save As → CSV (UTF-8). Use the same filename for the CSV.", False),
        ("5. Send the CSV + the same images folder back to the operator.", False),
        ("", False),
        ("Field length guidelines (AMA)", True),
        ("", False),
        ("• context_bar    < 8 words   — short hook on the black bar", False),
        ("• question_text  < 12 words  — community question, white pill", False),
        ("• answer_text    < 30 words  — the testimonial line, bubble", False),
        ("• image_note     optional    — used as AI outpaint prompt only", False),
    ]
    for r, (text, bold) in enumerate(notes_lines, start=1):
        cell = notes.cell(row=r, column=1, value=text)
        if bold:
            cell.font = Font(bold=True, size=12)
    notes.column_dimensions["A"].width = 80

    args.out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.out)
    print(f"[worksheet] wrote {args.out}  ({len(images)} rows)")

    # Clean up temp thumbnails (openpyxl reads them at save time)
    for p in tmp_paths:
        try: p.unlink()
        except Exception: pass

    return 0


def _natural_key(p: Path):
    """Sort v1, v2, …, v10 in numeric order instead of v1, v10, v2."""
    stem = p.stem.lower()
    digits = "".join(c for c in stem if c.isdigit())
    return (int(digits) if digits else 9_999, stem)


def _make_thumb(src: Path, size: int) -> Path:
    img = Image.open(src).convert("RGB")
    img.thumbnail((size, size), Image.LANCZOS)
    # Pad to a square so all rows line up the same height
    sq = Image.new("RGB", (size, size), (240, 240, 240))
    sq.paste(img, ((size - img.size[0]) // 2, (size - img.size[1]) // 2))
    tmp = Path(src.parent) / f".thumb_{src.stem}.png"
    sq.save(tmp, format="PNG")
    return tmp


def _load_prefill(path: Path, archetype: str) -> dict[str, dict[str, str]]:
    out: dict[str, dict[str, str]] = {}
    with open(path, newline="", encoding="utf-8-sig") as f:
        for row in csv_mod.DictReader(f):
            vid = (row.get("id") or "").strip()
            if not vid:
                continue
            out[vid] = {k: (row.get(k) or "").strip()
                        for k in ARCHETYPE_COLUMNS[archetype]
                        if k in row}
    return out


if __name__ == "__main__":
    sys.exit(main())
