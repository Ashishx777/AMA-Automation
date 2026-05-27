#!/usr/bin/env python3
"""prep_job.py — one-shot job-folder preparer.

Takes whatever the operator has — a folder of image files with random
names, optionally a CSV of copy in any column scheme — and emits a clean
job folder ready to drop into the pipeline:

    <out>/
        images/v1.<ext>, v2.<ext>, …      ← renamed copies, originals untouched
        copy.csv                          ← canonical schema, ids aligned to images
        worksheet.xlsx                    ← thumbnails + copy cells (pre-filled if CSV given)
        brand.json                        ← if --brand was passed

Usage::

    python prep_job.py \\
        --in-images /mnt/user-data/uploads/raw-pix \\
        --in-csv    /mnt/user-data/uploads/some-copy.csv   # optional
        --brand     mileenia                                # optional
        --out       /mnt/user-data/outputs/mileenia-job

Pairing rules:
- Images are sorted by a "natural" key (so `IMG_2.jpg < IMG_10.jpg`).
- CSV rows are paired positionally with sorted images:
    row 1 → v1, row 2 → v2, …
- If CSV has FEWER rows than images, extra images get blank copy cells.
- If CSV has MORE rows than images, extra rows get no `image_filename`.
- The original input CSV's column names (`context_bar/c1`, etc.) are
  normalized via the same synonym map the rest of the pipeline uses.
"""
from __future__ import annotations

import argparse
import csv as csv_mod
import json
import shutil
import sys
from pathlib import Path

_SCRIPT_DIR = Path(__file__).resolve().parent
sys.path.insert(0, str(_SCRIPT_DIR))

from utils import IMAGE_EXTS, read_csv, COLUMN_SYNONYMS


IMAGE_EXTS_LOWER = {e.lower() for e in IMAGE_EXTS}


def natural_key(p: Path):
    """Sort `v1, v2, …, v10` numerically rather than v1, v10, v2."""
    stem = p.stem.lower()
    digits = ""
    for ch in stem:
        if ch.isdigit():
            digits += ch
        elif digits:
            break
    return (int(digits) if digits else 9_999_999, stem)


def list_input_images(images_dir: Path) -> list[Path]:
    if not images_dir.exists():
        raise SystemExit(f"input images folder not found: {images_dir}")
    files = [p for p in images_dir.iterdir()
             if p.is_file() and p.suffix.lower() in IMAGE_EXTS_LOWER]
    if not files:
        raise SystemExit(f"no images found in {images_dir}")
    files.sort(key=natural_key)
    return files


def copy_and_rename(images: list[Path], out_images_dir: Path) -> list[dict]:
    """Copy each input image into out_images_dir as v1.<ext>, v2.<ext>, …
    Returns a list of dicts with {id, image_filename, original_name}."""
    out_images_dir.mkdir(parents=True, exist_ok=True)
    paired: list[dict] = []
    for i, src in enumerate(images, start=1):
        vid = f"v{i}"
        ext = src.suffix.lower()
        dst = out_images_dir / f"{vid}{ext}"
        shutil.copyfile(src, dst)
        paired.append({
            "id": vid,
            "image_filename": dst.name,
            "original_name": src.name,
        })
    return paired


def align_copy(rows: list[dict], paired_images: list[dict], archetype: str) -> list[dict]:
    """Position-pair the canonical CSV rows with the renamed images.

    Each output row gets:
      id, image_filename, source_image (blank), <archetype copy columns>, image_note
    """
    canonical_cols = COLUMN_SYNONYMS[archetype]
    out_rows: list[dict] = []
    n_images = len(paired_images)
    n_rows = len(rows)
    n = max(n_images, n_rows)

    for i in range(n):
        out = {"id": f"v{i + 1}"}
        # image_filename + source_image come from the paired images
        if i < n_images:
            out["image_filename"] = paired_images[i]["image_filename"]
            # source_image is operator-facing; for the auto-pair case we
            # surface the original name so they can sanity-check pairing.
            out["source_image"] = paired_images[i]["original_name"]
        else:
            out["image_filename"] = ""
            out["source_image"] = ""

        # copy columns from input CSV if present
        if i < n_rows:
            src_row = rows[i]
            for col in canonical_cols:
                if col == "id":
                    continue
                out[col] = src_row.get(col, "")
        else:
            for col in canonical_cols:
                if col == "id":
                    continue
                out[col] = ""

        out_rows.append(out)

    return out_rows


# Map canonical short keys to operator-friendly long header names,
# per archetype. The pipeline's synonym map (utils.COLUMN_SYNONYMS)
# accepts either, so emitting the long names doesn't break anything.
OUTPUT_HEADER_NAMES: dict[str, dict[str, str]] = {
    "ama": {
        "id": "id",
        "c1": "context_bar",
        "c2": "question_text",
        "c3": "answer_text",
        "image_note": "image_note",
    },
    "forum": {  # passthrough; forum already uses long names
        k: k for k in COLUMN_SYNONYMS["forum"]
    },
}


def write_csv(rows: list[dict], out_path: Path, archetype: str) -> Path:
    canonical = COLUMN_SYNONYMS[archetype]
    header_map = OUTPUT_HEADER_NAMES.get(archetype, {k: k for k in canonical})
    # Output column order: id, image_filename, source_image, copy fields, image_note
    canonical_keys = ["id"] + [c for c in canonical if c != "id"]
    output_headers = ["id", "image_filename", "source_image"] + [
        header_map.get(c, c) for c in canonical_keys if c != "id"
    ]
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with open(out_path, "w", encoding="utf-8", newline="") as f:
        w = csv_mod.writer(f)
        w.writerow(output_headers)
        for r in rows:
            row_values = [r.get("id", ""), r.get("image_filename", ""), r.get("source_image", "")]
            for c in canonical_keys:
                if c == "id":
                    continue
                row_values.append(r.get(c, ""))
            w.writerow(row_values)
    return out_path


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--in-images", required=True, type=Path,
                    help="Folder of raw image files (any names)")
    ap.add_argument("--in-csv", type=Path, default=None,
                    help="Optional CSV with copy. Column names normalized via synonym map.")
    ap.add_argument("--archetype", default="ama", choices=list(COLUMN_SYNONYMS))
    ap.add_argument("--brand", default=None, help="Brand id; writes brand.json if seeded")
    ap.add_argument("--out", required=True, type=Path,
                    help="Output job folder (created if missing)")
    args = ap.parse_args()

    args.out.mkdir(parents=True, exist_ok=True)

    # 1. Discover + rename images
    input_images = list_input_images(args.in_images)
    paired = copy_and_rename(input_images, args.out / "images")

    # 2. Read input CSV if any, normalize via synonym map
    rows: list[dict] = []
    if args.in_csv and args.in_csv.exists():
        rows = read_csv(args.in_csv, args.archetype)

    # 3. Align rows to images, write canonical copy.csv
    out_rows = align_copy(rows, paired, args.archetype)
    csv_out = write_csv(out_rows, args.out / "copy.csv", args.archetype)

    # 4. brand.json — seeded brand defaults or just an id
    brand_info = None
    if args.brand:
        defaults = {
            "rara":       {"name": "Rara",            "font": "system",  "accent": "#1a1612", "tier": "core"},
            "solvedskin": {"name": "The Solved Skin", "font": "elegant", "accent": "#2a1a14", "tier": "premium"},
            "mileenia":   {"name": "Mileenia",        "font": "elegant", "accent": "#c89a5e", "tier": "premium"},
            "zauv":       {"name": "Zauv",            "font": "display", "accent": "#1e1e1e", "tier": "core"},
            "parman":     {"name": "Parman",          "font": "serif",   "accent": "#2a2018", "tier": "premium"},
            "pragyanam":  {"name": "Pragyanam",       "font": "serif",   "accent": "#2d2014", "tier": "premium"},
            "gyros":      {"name": "Gyros",           "font": "rounded", "accent": "#1d1d1d", "tier": "core"},
            "subtle":     {"name": "Subtle",          "font": "inter",   "accent": "#1a1a1a", "tier": "core"},
            "atovio":     {"name": "Atovio",          "font": "display", "accent": "#111111", "tier": "core"},
        }
        brand_info = {"id": args.brand}
        brand_info.update(defaults.get(args.brand, {"name": args.brand, "font": "system", "accent": "#111111"}))
        brand_info["paddingStyle"] = "solid"
        (args.out / "brand.json").write_text(json.dumps(brand_info, indent=2), encoding="utf-8")

    # 5. Worksheet.xlsx — delegate to build_worksheet.py, using the new
    # copy.csv as pre-fill so the cells aren't blank if copy was provided.
    worksheet_path = args.out / "worksheet.xlsx"
    try:
        # Import lazily so prep can be used standalone if openpyxl is missing.
        from build_worksheet import _make_thumb, ARCHETYPE_COLUMNS, _load_prefill  # type: ignore
        from openpyxl import Workbook
        from openpyxl.drawing.image import Image as XLImage
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "Copy worksheet"

        columns = ["image", "id", "image_filename"] + ARCHETYPE_COLUMNS[args.archetype]
        for col_idx, name in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=name)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1f1f23")
            cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 28
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 8
        ws.column_dimensions["C"].width = 22
        for i in range(len(ARCHETYPE_COLUMNS[args.archetype])):
            ws.column_dimensions[get_column_letter(4 + i)].width = 40

        prefill = {r["id"]: r for r in out_rows}
        tmp_thumbs: list[Path] = []
        for i, pi in enumerate(paired, start=1):
            row_num = i + 1
            src = args.out / "images" / pi["image_filename"]
            thumb = _make_thumb(src, 200)
            tmp_thumbs.append(thumb)
            xlimg = XLImage(str(thumb)); xlimg.width = 200; xlimg.height = 200
            ws.add_image(xlimg, f"A{row_num}")
            ws.row_dimensions[row_num].height = 150
            ws.cell(row=row_num, column=2, value=pi["id"]).alignment = Alignment(vertical="center")
            ws.cell(row=row_num, column=3, value=pi["image_filename"]).alignment = Alignment(vertical="center", wrap_text=True)
            row_data = prefill.get(pi["id"], {})
            for j, col in enumerate(ARCHETYPE_COLUMNS[args.archetype]):
                ws.cell(row=row_num, column=4 + j, value=row_data.get(col, "")).alignment = Alignment(vertical="center", wrap_text=True)
        ws.freeze_panes = "A2"

        notes = wb.create_sheet("Read me")
        for r, txt in enumerate([
            "Auto-generated worksheet — prepared by prep_job.py",
            "",
            "Don't touch columns A (image), B (id), C (image_filename).",
            f"Fill: {', '.join(ARCHETYPE_COLUMNS[args.archetype])}.",
            "Save As CSV (UTF-8) and send back, or just drop this whole folder back to claude.ai.",
        ], start=1):
            notes.cell(row=r, column=1, value=txt)
        notes.column_dimensions["A"].width = 80

        wb.save(worksheet_path)
        for t in tmp_thumbs:
            try: t.unlink()
            except Exception: pass
        worksheet_status = "written"
    except Exception as e:
        worksheet_status = f"skipped ({e})"

    # Summary
    summary = {
        "out": str(args.out),
        "images_count": len(paired),
        "csv_rows_count": len(rows),
        "csv_rows_with_image": min(len(rows), len(paired)),
        "images_with_no_copy": max(0, len(paired) - len(rows)),
        "copy_rows_with_no_image": max(0, len(rows) - len(paired)),
        "brand": brand_info["id"] if brand_info else None,
        "worksheet": worksheet_status,
        "rename_map": [{"from": pi["original_name"], "to": pi["image_filename"]} for pi in paired],
    }
    print("[prep-summary] " + json.dumps(summary))
    return 0


if __name__ == "__main__":
    sys.exit(main())
